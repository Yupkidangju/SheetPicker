import os
import pandas as pd
from pathlib import Path
from typing import List, Generator, Dict, Any
from openpyxl import load_workbook
from src.utils.logger import logger

class FileScanner:
    """
    [KR] 파일 시스템 스캐너 및 데이터 로더 클래스.
    지정된 경로에서 엑셀 및 CSV 파일을 탐색하고, 대용량 데이터를 안전하게 로드합니다.
    """

    SUPPORTED_EXTENSIONS = {'.xlsx', '.xls', '.csv'}

    def get_supported_files(self, paths: List[str]) -> List[str]:
        """
        [KR] 입력된 경로 리스트(파일 또는 폴더)에서 지원되는 포맷의 파일들을 찾아 반환합니다.
        폴더의 경우 재귀적으로 탐색합니다.

        Args:
            paths (List[str]): 검색할 파일 또는 폴더 경로 리스트

        Returns:
            List[str]: 발견된 파일의 절대 경로 리스트
        """
        found_files = []

        for p in paths:
            # os.path 사용으로 성능 최적화 (pathlib.Path 객체 생성 오버헤드 제거)
            p_abs = os.path.abspath(p)
            if not os.path.exists(p_abs):
                continue

            if os.path.isfile(p_abs):
                ext = os.path.splitext(p_abs)[1].lower()
                if ext in self.SUPPORTED_EXTENSIONS:
                    found_files.append(p_abs)
            elif os.path.isdir(p_abs):
                for root, _, files in os.walk(p_abs):
                    for file in files:
                        ext = os.path.splitext(file)[1].lower()
                        if ext in self.SUPPORTED_EXTENSIONS:
                            found_files.append(os.path.join(root, file))

        # [KR] 중복 제거 후 반환
        return list(set(found_files))

    def read_file_chunks(self, file_path: str, chunksize: int = 10000) -> Generator[Dict[str, Any], None, None]:
        """
        [KR] 파일을 지정된 chunksize만큼 나누어 읽어오는 제너레이터입니다.
        대용량 파일 처리 시 메모리 사용을 최소화하기 위해 사용합니다.

        Args:
            file_path (str): 읽을 파일의 경로
            chunksize (int): 한 번에 읽을 행(Row)의 수

        Yields:
            Dict[str, Any]: {'sheet_name': str, 'data': pd.DataFrame} 형태의 딕셔너리
            에러 발생 시 예외가 전파됩니다.
        """
        file_path_obj = Path(file_path)
        ext = file_path_obj.suffix.lower()

        # [KR] 파일 존재 확인
        if not file_path_obj.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        try:
            if ext == '.csv':
                # [KR] CSV 파일 로드 (UTF-8-SIG, 에러 라인 스킵)
                sheet_name = file_path_obj.name
                reader = pd.read_csv(
                    file_path,
                    chunksize=chunksize,
                    encoding='utf-8-sig',
                    on_bad_lines='skip'
                )
                for chunk in reader:
                    yield {'sheet_name': sheet_name, 'data': chunk}

            elif ext == '.xlsx':
                # [KR] .xlsx 파일: openpyxl read_only 모드로 메모리 효율적 로딩
                wb = None
                try:
                    wb = load_workbook(file_path, read_only=True, data_only=True)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        rows_iter = ws.iter_rows(values_only=True)

                        # 헤더 읽기
                        try:
                            header = next(rows_iter)
                        except StopIteration:
                            continue # 빈 시트 스킵

                        buffer = []
                        for row in rows_iter:
                            buffer.append(row)
                            if len(buffer) >= chunksize:
                                df_chunk = pd.DataFrame(buffer, columns=header)
                                yield {'sheet_name': sheet_name, 'data': df_chunk}
                                buffer = []

                        # 남은 데이터 처리
                        if buffer:
                            df_chunk = pd.DataFrame(buffer, columns=header)
                            yield {'sheet_name': sheet_name, 'data': df_chunk}
                finally:
                    # [KR] 제너레이터 중단 시에도 파일 리소스 해제 보장
                    if wb:
                        wb.close()

            elif ext == '.xls':
                # [KR] .xls 파일: 레거시 포맷은 pandas 기본 로드 (전체 로드 후 청크 분할)
                xls = pd.ExcelFile(file_path)
                try:
                    for sheet_name in xls.sheet_names:
                        # [KR] 파일 경로 대신 ExcelFile 객체를 사용하여 I/O 오버헤드 감소
                        df = pd.read_excel(xls, sheet_name=sheet_name)
                        # DataFrame을 chunksize로 슬라이싱하여 yield
                        num_rows = len(df)
                        for i in range(0, num_rows, chunksize):
                            yield {'sheet_name': sheet_name, 'data': df.iloc[i:i+chunksize]}
                finally:
                    xls.close()

        except Exception as e:
            # [KR] 읽기 실패 시 로깅 후 예외 전파
            logger.error(f"Failed to read file {file_path}: {e}")
            raise RuntimeError(f"Failed to read file {file_path}: {e}")
